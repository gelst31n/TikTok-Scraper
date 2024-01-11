import pandas as pd
import re
import math
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment
from openpyxl.utils import column_index_from_string

# Patterns for Instagram and Email
instagram_pattern = r'\b(?:ig(?:\.|\s*@\s*)?|inst(?:a(?:gram)?)?\s*:\s*|\s*📸)\s*([\w\.]+)'
email_pattern = r'[\w\.-]+@[\w\.-]+'

data = pd.read_csv('sins OUT NOW.csv')
data = data.filter(regex='^(?!Unnamed)')
data = data.drop(columns='secUid')
data.dropna(inplace=True)
data = data[data['median views in last 28 days'] <= 100000]
data = data[data['posts in last 28 days'] >= 7]
data = data[data['median views per follower'] >= 0.25]
data = data.loc[(data['median views in last 28 days'] >= 1000) | (data['mean views in last 28 days'] >= 10000)]

data['median likes per 100 views'] = data['median likes in last 28 days'] / data['median views in last 28 days'] * 100
data['median comments per 1000 views'] = data['median comments in last 28 days'] / data[
    'median views in last 28 days'] * 1000

data = data.drop(columns='median likes in last 28 days')
data = data.drop(columns='median comments in last 28 days')

# Extracting Instagram handles
data['instagram'] = data['bio'].apply(lambda x: ' '.join(re.findall(instagram_pattern, x, re.IGNORECASE)))
data['instagram_link'] = data['instagram'].apply(lambda x: f'https://www.instagram.com/{x}' if x else '')

# Extracting emails
data['email'] = data['bio'].apply(lambda x: ' '.join(re.findall(email_pattern, x)))

cols = list(data.columns)
cols.remove('tags')
cols.append('tags')
data = data[cols]
data['tags'] = data['tags'].astype(str)
data['tags'] = (
    data['tags']
    .str.replace('[', '', regex=False)
    .str.replace(']', '', regex=False)
    .str.replace("'", '', regex=False)
    .str.replace('"', '', regex=False)
)

data = pd.concat([data[data.columns.difference(['bio'], sort=False)], data[['bio']]], axis=1)

# Save the dataframe to an excel file without the 'instagram_link' column
filename = "sins OUT NOW.xlsx"
data.drop(columns=['link', 'instagram_link']).to_excel(filename, index=False, engine='openpyxl')

# Open the excel file and add hyperlinks
book = load_workbook(filename)
worksheet = book.active


# Convert column number to Excel-style column letter
def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


# Determine the column letter of the 'instagram' column
tt_col = colnum_string(data.columns.get_loc('username') + 1)
ig_col = colnum_string(data.columns.get_loc('instagram'))
followers_col = colnum_string(data.columns.get_loc('followers'))
views_col = colnum_string(data.columns.get_loc('median views in last 28 days'))
mean_views_col = colnum_string(data.columns.get_loc('mean views in last 28 days'))
likes_col = colnum_string(data.columns.get_loc('median likes per 100 views'))
comments_col = colnum_string(data.columns.get_loc('median comments per 1000 views'))
view_ratio_col = colnum_string(data.columns.get_loc('median views per follower'))

# Apply hyperlinks
for idx, link in enumerate(data['instagram_link'],
                           start=2):  # start=2 because Excel uses 1-indexing and we skip the header
    cell = f'{ig_col}{idx}'
    worksheet[cell].hyperlink = link
    worksheet[cell].style = "Hyperlink"  # optional: to make it look like a hyperlink

# Apply hyperlinks
for idx, link in enumerate(data['link'], start=2):  # start=2 because Excel uses 1-indexing and we skip the header
    cell = f'{tt_col}{idx}'
    worksheet[cell].hyperlink = link
    worksheet[cell].style = "Hyperlink"  # optional: to make it look like a hyperlink

# Define the number format style with commas
comma_style = NamedStyle(name='comma_style', number_format='#,##0')
decimal_style = NamedStyle(name='decimal_style', number_format='0.00')

# Apply the style to each cell in the 'Number' column (Column B in Excel)
for row in range(2, worksheet.max_row + 1):  # Starting from row 2 to skip the header
    for col in [views_col, followers_col, mean_views_col]:
        cell = worksheet[f"{col}{row}"]  # Combine column letter and row number to get the cell reference
        cell.style = comma_style

# Apply the style to each cell in the 'Number' column (Column B in Excel).
for row in range(2, worksheet.max_row + 1):  # Starting from row 2 to skip the header
    for col in [likes_col, comments_col, view_ratio_col]:
        cell = worksheet[f"{col}{row}"]  # Combine column letter and row number to get the cell reference
        cell.style = decimal_style

for i in ['median likes per 100 views', 'median comments per 1000 views', 'median views per follower']:
    target_col = colnum_string(data.columns.get_loc(i))
    # Convert column letter to index
    target_col_idx = column_index_from_string(target_col)

    # Get min and max values from the column for normalization
    min_value = min(worksheet.cell(row=row, column=target_col_idx).value for row in range(2, worksheet.max_row + 1))
    max_value = max(worksheet.cell(row=row, column=target_col_idx).value for row in range(2, worksheet.max_row + 1))

    min_value_log = math.log1p(min_value)  # log1p is used to handle the case when the value is zero
    max_value_log = math.log1p(max_value)


    def make_color_paler(red, green, blue, alpha=0.5):
        white = 255
        new_red = int((1 - alpha) * red + alpha * white)
        new_green = int((1 - alpha) * green + alpha * white)
        new_blue = int((1 - alpha) * blue + alpha * white)
        return new_red, new_green, new_blue


    # Loop through each cell in the column to color it
    for row in range(2, worksheet.max_row + 1):  # Starting from row 2 to skip the header
        cell = worksheet.cell(row=row, column=target_col_idx)
        value = cell.value

        # Logarithmically scale and normalize the cell value
        value_log = math.log1p(value)
        normalized_value = (value_log - min_value_log) / (max_value_log - min_value_log)

        # Calculate RGB components. Transition from red to yellow to green.
        if normalized_value < 0.5:
            red = 255
            green = int(255 * (normalized_value / 0.5))
            blue = 0
        else:
            red = int(255 * ((1.0 - normalized_value) / 0.5))
            green = 255
            blue = 0

        # Make the color paler
        red, green, blue = make_color_paler(red, green, blue, alpha=0.5)

        # Convert RGB to Hex
        color_hex = "{:02X}{:02X}{:02X}".format(red, green, blue)

        # Create fill with the calculated color
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

# Rename 'median views' to 'views'
for col_idx, column in enumerate(worksheet.iter_cols(values_only=True, max_row=1), 1):
    for cell_value in column:
        if cell_value == 'posts in last 28 days':
            worksheet.cell(row=1, column=col_idx).value = 'posts \n (past 28 days)'
        if cell_value == 'median views in last 28 days':
            worksheet.cell(row=1, column=col_idx).value = 'median views \n (past 28 days)'
        if cell_value == 'mean views in last 28 days':
            worksheet.cell(row=1, column=col_idx).value = 'mean views \n (past 28 days)'
        if cell_value == 'median views per follower':
            worksheet.cell(row=1, column=col_idx).value = 'views per follower \n (past 28 days)'
        if cell_value == 'median likes per 100 views':
            worksheet.cell(row=1, column=col_idx).value = 'likes per 100 views \n (past 28 days)'
        if cell_value == 'median comments per 1000 views':
            worksheet.cell(row=1, column=col_idx).value = 'comments per 1k views \n (past 28 days)'

columns_to_adjust = ['A', 'B', 'C', 'D', 'K', 'L']
columns_2 = ['E', 'F', 'G', 'H', 'I', 'J']  # Add as many columns as you want

for col_letter in columns_to_adjust:
    column = worksheet[col_letter]
    max_length = 0
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    worksheet.column_dimensions[col_letter].width = max_length + 2

for col_letter in columns_2:
    column = worksheet[col_letter]
    max_length = 0
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    worksheet.column_dimensions[col_letter].width = max_length - 7

for row in worksheet.iter_rows():
    for cell in row:
        cell.font = Font(name='Arial')
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Capitalize all column titles in the first row
for cell in worksheet["1:1"]:
    cell.value = cell.value.title() if cell.value else cell.value
    cell.font = Font(name='Arial', bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

worksheet.auto_filter.ref = worksheet.dimensions

# Save the changes
book.save(filename)
